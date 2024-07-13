from django.shortcuts import render,redirect
from calendar import HTMLCalendar
from datetime import datetime
import calendar
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, redirect
from django.core.mail import send_mail,EmailMessage
from django.conf import settings
import mimetypes
from . forms import RegisterUserForm, EnregistrerChargeForm,EnregistrerFormulaireChargeForm,EnregistrerFormulaireListeProprietaireForm,EnregistrerFormulaireCotizationForm,EnregistrerFormulairePConciergeForm
from .models import Charge,FormulaireCharge,FormulaireListeDesProprietaires,FormulaireCotization,FormulairePConcierge
from django.db.models import Sum
from django.http import JsonResponse
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
import io
from django.http import FileResponse
from django.urls import reverse
import os
import tempfile
import csv
from django.http import HttpResponse
from excel_response import ExcelResponse  # Importa la respuesta Excel
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from datetime import date
from openpyxl.styles import Font
from django.core import management

#import datetime
# Create your views here.


def home(request, year=datetime.now().year, month=datetime.now().strftime('%B')):
    name = "John"
    name = request.user.username
    month = month.capitalize()
    day = datetime.today()
    
    # convert month name to number
    month_number = list(calendar.month_name).index(month)
    month_number = int(month_number)
       
    # create calendar
    cal = calendar.HTMLCalendar().formatmonth(year, month_number)
    # create current year
    now = datetime.now()
    current_year = now.year

    # query event date
        
    time = now.strftime('%I:%M:%S %p')

    context = {"name": name, "year": year, "month": month, "month_number": month_number, "cal": cal, "current_year": current_year, "time": time, "day":day}
    return render(request, 'gestion_immeuble_app/home.html', context)


def login_user(request):
    
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
        # Redirect to a success page.
            return redirect('home')
        else:
            # Return an 'invalid login' error message.
            messages.success(request, "Erreur durant le login essaye une autre fois")
            return redirect('login')
    else:
        context = {}
        return render(request, 'gestion_immeuble_app/login.html', context)

def logout_user(request):
    logout(request)
    messages.success(request, "tu as été deconnecté")
    return redirect('home')

def register_user(request):
    
    if request.method == "POST":
        form=RegisterUserForm(request.POST)
        if form.is_valid():
            form.save()
            username=form.cleaned_data['username']
            password=form.cleaned_data['password1']
            
            user = authenticate(username=username, password=password)
            
            login(request,user)
            
            messages.success(request, "registracion successfull")
            return redirect('home')
    else:
        form=RegisterUserForm()
        
    context={'form':form}
    return render(request, 'gestion_immeuble_app/register_user.html', context)   


def enregistrer_charge_view(request):
    if request.method == 'POST':
        form = EnregistrerChargeForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            nome_charge = instance.nome_charge
           
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {nome_charge} - {nome_charge}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Nom de la charge: {nome_charge}\nNom de la charge: {nome_charge}\nNom de la charge: {nome_charge}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            #if serie_pelicula_imagen:
                #mime_type, _ = mimetypes.guess_type(serie_pelicula_imagen.path)
                #email_message.attach(serie_pelicula_imagen.name, serie_pelicula_imagen.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)

            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_des_charges')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerChargeForm()
    return render(request, 'enregistrer_charge.html', {'form': form})


def liste_des_charges(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_charges = Charge.objects.all()
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_charges, 5)
    page = request.GET.get('page')
    tous_les_charges = p.get_page(page)
    nums = "a" * tous_les_charges.paginator.num_pages
    
    print("hola : " + str(tous_les_charges.paginator.num_pages))
    
    context = {'la_lista_des_charges': la_lista_des_charges, 'tous_les_charges': tous_les_charges, 'nums': nums, 'name':name}
    return render(request, 'gestion_immeuble_app/la_liste_des_charges.html', context)

def actualiser_la_charge(request, id_charge):
    charge = Charge.objects.get(pk=id_charge)
    form = EnregistrerChargeForm(request.POST or None, request.FILES or None,  instance=charge)
    if form.is_valid():
        form.save()
        return redirect('liste_des_charges')
    context = {'charge': charge, 'form': form}
    return render(request, 'gestion_immeuble_app/actualizer_la_charge.html', context)

def eliminer_la_charge(request, id_charge):
    charge = get_object_or_404(Charge, id=id_charge)
    charge.delete()
    messages.success(request, "La charge a été eliminer correctement.")
    return redirect('liste_des_charges')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

def liste_des_formulaire_charges(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_formulaire_charges = FormulaireCharge.objects.all()
    total_montant = la_lista_des_formulaire_charges.aggregate(total=Sum('montant'))['total']
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_formulaire_charges, 5)
    page = request.GET.get('page')
    tous_les_formulaire_charges = p.get_page(page)
    nums = "a" * tous_les_formulaire_charges.paginator.num_pages
    
    print("hola : " + str(tous_les_formulaire_charges.paginator.num_pages))
    
    context = {'la_lista_des_formulaire_charges': la_lista_des_formulaire_charges, 'tous_les_formulaire_charges': tous_les_formulaire_charges, 'nums': nums, 'name':name, 'total_montant':total_montant}
    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = datetime.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulaire Concierge"
        
        # Crear una nueva hoja para la tabla
        #ws2 = wb.create_sheet(title="Formulaire Concierge")

        # Agregar los encabezados de la tabla
        headers = ['Date', 'Charge', 'Du', 'Au', 'Montant', 'Nom Fichier']
        #if request.user.is_superuser:
            #headers.extend(['Imprimer pdf', 'Actualiser', 'Eliminer'])
        ws.append(headers)

        # Aplicar negrita a los encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Agregar los datos de la tabla
        for item in la_lista_des_formulaire_charges:
            row = [
                item.date.strftime('%Y-%m-%d'),
                item.charge.nome_charge,
                item.du,
                item.au,
                item.montant,
                item.image_charge.name
            ]
            #if request.user.is_superuser:
                #row.extend(['Imprimer', 'Actualiser', 'Eliminer'])
            ws.append(row)
            
        ws.append(['Total Montant', '', '','', total_montant])
        
        # Crear una respuesta HTTP con el archivo Excel
        filename = f"liste_charges_immeuble_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=lista_paie_concierge.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    return render(request, 'gestion_immeuble_app/la_liste_des_formulaire_charges.html', context)


def enregistrer_formulaire_charge_view(request):
    if request.method == 'POST':
        form = EnregistrerFormulaireChargeForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            charge = instance.charge
            du = instance.du
            au = instance.au
            date = instance.date
            montant=instance.montant
            image_charge = instance.image_charge
           
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {date} - {charge} - {montant}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Nom de la charge: {charge}\nDate de la charge: {date}\nMontant de la charge: {montant}\nDu de la charge: {du}\nAu de la charge: {au}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            if image_charge:
                mime_type, _ = mimetypes.guess_type(image_charge.path)
                email_message.attach(image_charge.name, image_charge.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)

            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_des_formulaire_charges')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerFormulaireChargeForm()
    return render(request, 'enregistrer_formulaire_charge.html', {'form': form})


def actualiser_formulaire_charge(request, id_formulaire_charge):
    formulaire_charge = FormulaireCharge.objects.get(pk=id_formulaire_charge)
    form = EnregistrerFormulaireChargeForm(request.POST or None, request.FILES or None,  instance=formulaire_charge)
    if form.is_valid():
        form.save()
        return redirect('liste_des_formulaire_charges')
    context = {'formulaire_charge': formulaire_charge, 'form': form}
    return render(request, 'gestion_immeuble_app/actualizer_formulaire_charge.html', context)

def eliminer_formulaire_charge(request, id_formulaire_charge):
    formulaire_charge = get_object_or_404(FormulaireCharge, id=id_formulaire_charge)
    formulaire_charge.delete()
    messages.success(request, "Le formulaire charge a été eliminer correctement.")
    return redirect('liste_des_formulaire_charges')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal


def liste_des_formulaire_liste_proprietaires(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_formulaire_liste_proprietaires = FormulaireListeDesProprietaires.objects.all()
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_formulaire_liste_proprietaires, 5)
    page = request.GET.get('page')
    tous_les_formulaire_liste_proprietaires = p.get_page(page)
    nums = "a" * tous_les_formulaire_liste_proprietaires.paginator.num_pages
    
    print("hola : " + str(tous_les_formulaire_liste_proprietaires.paginator.num_pages))
    
    context = {'la_lista_des_formulaire_liste_proprietaires': la_lista_des_formulaire_liste_proprietaires, 'tous_les_formulaire_liste_proprietaires': tous_les_formulaire_liste_proprietaires, 'nums': nums, 'name':name}
    return render(request, 'gestion_immeuble_app/la_lista_des_formulaire_liste_proprietaires.html', context)


def enregistrer_formulaire_liste_proprietaire_view(request):
    if request.method == 'POST':
        form = EnregistrerFormulaireListeProprietaireForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            nom = instance.nom
            prenom = instance.prenom
            aptNum = instance.aptNum
            tel = instance.tel
            email=instance.email
            proprietaire_locataire = instance.proprietaire_locataire
           
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {nom} - {prenom} - {proprietaire_locataire}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Nom : {nom}\nPrenom: {prenom}\nNº Apt: {aptNum}\nTel: {tel}\nEmail: {email}\nProprietaire/Locataire: {proprietaire_locataire}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            #if image_charge:
                #mime_type, _ = mimetypes.guess_type(image_charge.path)
                #email_message.attach(image_charge.name, image_charge.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)

            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_des_formulaire_liste_proprietaires')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerFormulaireListeProprietaireForm()
    return render(request, 'enregistrer_formulaire_liste_proprietaire.html', {'form': form})


def actualiser_formulaire_liste_proprietaires(request, id_formulaire_liste_proprietaire):
    formulaire_liste_proprietaires = FormulaireListeDesProprietaires.objects.get(pk=id_formulaire_liste_proprietaire)
    form = EnregistrerFormulaireListeProprietaireForm(request.POST or None, request.FILES or None,  instance=formulaire_liste_proprietaires)
    if form.is_valid():
        form.save()
        return redirect('liste_des_formulaire_liste_proprietaires')
    context = {'formulaire_liste_proprietaires': formulaire_liste_proprietaires, 'form': form}
    return render(request, 'gestion_immeuble_app/actualizer_formulaire_liste_proprietaire.html', context)

def eliminer_formulaire_liste_proprietaires(request, id_formulaire_liste_proprietaire):
    formulaire_liste_proprietaire = get_object_or_404(FormulaireListeDesProprietaires, id=id_formulaire_liste_proprietaire)
    formulaire_liste_proprietaire.delete()
    messages.success(request, "Le formulaire liste des proprietaires a été eliminer correctement.")
    return redirect('liste_des_formulaire_liste_proprietaires')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

def liste_formulaire_cotization(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_formulaire_cotization = FormulaireCotization.objects.all()
    total_montant = la_lista_formulaire_cotization.aggregate(total=Sum('montant'))['total']
  
    # Agrupar y sumar Montant por combinación de Nom y Prenom
    montant_por_usuario = {}
    for formulaire_cotization in la_lista_formulaire_cotization:
        key = (formulaire_cotization.nom, formulaire_cotization.prenom)
        if key in montant_por_usuario:
            montant_por_usuario[key] += formulaire_cotization.montant
        else:
            montant_por_usuario[key] = formulaire_cotization.montant
    
     # Convertir el diccionario en una lista de tuplas para usar en la plantilla
    montant_por_usuario_list = [{'nom': key[0], 'prenom': key[1], 'total_montant': value} for key, value in montant_por_usuario.items()]
    
    
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_formulaire_cotization, 5)
    page = request.GET.get('page')
    tous_les_formulaire_cotization = p.get_page(page)
    nums = "a" * tous_les_formulaire_cotization.paginator.num_pages
    
    print("hola : " + str(tous_les_formulaire_cotization.paginator.num_pages))
    
    context = {'la_lista_formulaire_cotization': la_lista_formulaire_cotization, 'tous_les_formulaire_cotization': tous_les_formulaire_cotization, 'nums': nums, 'name':name, 'total_montant': total_montant,'montant_por_usuario': montant_por_usuario_list,}
    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = datetime.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulaire cotisation"
        
        # Crear una nueva hoja para la tabla
        #ws2 = wb.create_sheet(title="Formulaire Concierge")

        # Agregar los encabezados de la tabla
        headers = ['Nº Ordre', 'Date', 'Nom', 'Prenom', 'Nº Apt', 'Montant','Reglé?','Mois', 'Année' , 'Frais Sindic', 'Frais Sindic Manual']
        #if request.user.is_superuser:
            #headers.extend(['Imprimer pdf', 'Actualiser', 'Eliminer'])
        ws.append(headers)

        # Aplicar negrita a los encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Agregar los datos de la tabla
        for item in la_lista_formulaire_cotization:
            row = [
                item.codeFormCotiz,
                item.date.strftime('%Y-%m-%d'),
                item.nom,
                item.prenom,
                item.aptNum,
                item.montant,
                item.regle,
                item.motif_mois,
                item.motif_annee,
                item.frais_sindic,
                item.frais_sindic_manual
            ]
            #if request.user.is_superuser:
                #row.extend(['Imprimer', 'Actualiser', 'Eliminer'])
            ws.append(row)
        
        # Agregar los datos agrupados por usuario
        for montant in montant_por_usuario_list:
            row = [
                'Total Montant par Proprietaire/Locataire',
                montant['nom'],
                montant['prenom'],
                montant['total_montant']
            ]
            ws.append(row)
        ws.append(['Total Montant', '', '', total_montant])
        
        # Crear una respuesta HTTP con el archivo Excel
        filename = f"liste_des_cotisations_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=lista_paie_concierge.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    
    # Agrupar montants por usuario (nom, prenom)
    montant_por_usuario1 = {}
    for formulaire_cotization in la_lista_formulaire_cotization:
        key = (formulaire_cotization.nom, formulaire_cotization.prenom)
        if key in montant_por_usuario1:
            montant_por_usuario1[key].append(formulaire_cotization)
        else:
            montant_por_usuario1[key] = [formulaire_cotization]
    
    # Convertir el diccionario en una lista de diccionarios para usar en la plantilla
    montant_por_usuario_list1 = [{'nom': key[0], 'prenom': key[1], 'total_montant1': sum(fc.montant for fc in value)} for key, value in montant_por_usuario1.items()]

    
    # Si se solicita la exportación a Excel
    if request.GET.get('export') == 'excel1':
        nom = request.GET.get('nom')
        prenom = request.GET.get('prenom')
        
        # Obtener los formulario_cotizations para el nom y prenom especificados
        if nom and prenom:
            key = (nom, prenom)
            formulario_cotizations = montant_por_usuario1.get(key, [])
            
            total_montant_usuario = sum(fc.montant for fc in formulario_cotizations)
            
            # Crear un libro de trabajo y una hoja
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Form cot_{nom}_{prenom}"
            
            # Agregar los encabezados de la tabla
            headers = ['Nº Ordre', 'Date', 'Nom', 'Prenom', 'Nº Apt', 'Montant', 'Reglé?', 'Mois', 'Année', 'Frais Sindic', 'Frais Sindic Manual']
            ws.append(headers)
            
            # Aplicar negrita a los encabezados
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            # Agregar los datos para el nom y prenom especificados
            for formulaire_cotization in formulario_cotizations:
                row = [
                    formulaire_cotization.codeFormCotiz,
                    formulaire_cotization.date.strftime('%Y-%m-%d'),
                    formulaire_cotization.nom,
                    formulaire_cotization.prenom,
                    formulaire_cotization.aptNum,
                    formulaire_cotization.montant,
                    formulaire_cotization.regle,
                    formulaire_cotization.motif_mois,
                    formulaire_cotization.motif_annee,
                    formulaire_cotization.frais_sindic,
                    formulaire_cotization.frais_sindic_manual
                ]
                ws.append(row)
            ws.append(['Total Montant', '', '','','', total_montant_usuario])
            
            # Crear una respuesta HTTP con el archivo Excel para descarga
            filename = f"liste_des_cotisations_{nom}_{prenom}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
    
    
    return render(request, 'gestion_immeuble_app/la_lista_formulaire_cotization.html', context)

def get_proprietaire_data(request, pk):
    try:
        proprietaire = FormulaireListeDesProprietaires.objects.get(pk=pk)
        data = {
            'nom': proprietaire.nom,
            'prenom': proprietaire.prenom,
            'aptNum': proprietaire.aptNum,
        }
        return JsonResponse(data)
    except FormulaireListeDesProprietaires.DoesNotExist:
        return JsonResponse({'error': 'Proprietaire not found'}, status=404)

def enregistrer_formulaire_cotization_view(request):
    if request.method == 'POST':
        form = EnregistrerFormulaireCotizationForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            nom = instance.nom
            prenom = instance.prenom
            aptNum = instance.aptNum
            date = instance.date
            montant=instance.montant
            regle = instance.regle
            motif_mois = instance.motif_mois
            motif_annee = instance.motif_annee
            frais_sindic = instance.frais_sindic
            liste_proprietaire = instance.liste_proprietaire
           
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {nom} - {prenom} - {montant}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Nom : {nom}\nPrenom: {prenom}\nNº Apt: {aptNum}\nDate: {date}\nMotif mois: {motif_mois}\nMotif année: {motif_annee}\nFrais Sindic: {frais_sindic}\nListe Proprietaire: {liste_proprietaire}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            #if image_charge:
                #mime_type, _ = mimetypes.guess_type(image_charge.path)
                #email_message.attach(image_charge.name, image_charge.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)

            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_formulaire_cotization')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerFormulaireCotizationForm()
    return render(request, 'enregistrer_formulaire_cotization.html', {'form': form})


def actualiser_formulaire_cotization(request, id_formulaire_cotization):
    formulaire_cotization = FormulaireCotization.objects.get(pk=id_formulaire_cotization)
    form = EnregistrerFormulaireCotizationForm(request.POST or None, request.FILES or None,  instance=formulaire_cotization)
    if form.is_valid():
        form.save()
        return redirect('liste_formulaire_cotization')
    context = {'formulaire_cotization': formulaire_cotization, 'form': form}
    return render(request, 'gestion_immeuble_app/actualizer_formulaire_cotization.html', context)

def eliminer_formulaire_cotization(request, id_formulaire_cotization):
    formulaire_cotization = get_object_or_404(FormulaireCotization, id=id_formulaire_cotization)
    formulaire_cotization.delete()
    messages.success(request, "Le formulaire cotization a été eliminer correctement.")
    return redirect('liste_formulaire_cotization')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

def get_last_id(request):
    last_instance = FormulaireCotization.objects.all().order_by('id').last()
    last_id = last_instance.id if last_instance else 0
    return JsonResponse({'last_id': last_id})


def liste_situation_caisse(request):
    # venue_list = Venue.objects.all().order_by('?')
    #la_lista_formulaire_cotization = FormulaireCotization.objects.all()
    la_lista_formulaire_cotization = FormulaireCotization.objects.filter(regle=True)
    total_montant_cotizacion = la_lista_formulaire_cotization.aggregate(total=Sum('montant'))['total']
    
    la_lista_des_formulaire_charges = FormulaireCharge.objects.all()
    total_montant_charge = la_lista_des_formulaire_charges.aggregate(total=Sum('montant'))['total']
    if total_montant_cotizacion is None  :
        total_montant_cotizacion = 0
    
    if total_montant_charge is None  :
        total_montant_charge = 0
    
    #total =  total_montant_cotizacion - total_montant_charge   
    # Calculating total only if both values are not None
    if total_montant_cotizacion is not None and total_montant_charge is not None :
        total = total_montant_cotizacion - total_montant_charge
    else:
        total = 0  # Default to 0 if either value is None
    name = request.user.username
    
    
    
    context = {'la_lista_formulaire_cotization': la_lista_formulaire_cotization, 'la_lista_des_formulaire_charges': la_lista_des_formulaire_charges, 'name':name, 'total_montant_cotizacion': total_montant_cotizacion,'total_montant_charge': total_montant_charge, 'total' : total,}
    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = date.today()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Situation Caisse"
        
        ws.append(['Date', today.strftime('%Y-%m-%d')])
        # Agregar los datos a la hoja
        ws.append(['Montant Charge', total_montant_charge, " dh"])
        ws.append(['Montant Cotisation', total_montant_cotizacion , " dh"])
        ws.append(['Total ( Cotisation - Charge )', total , " dh"])
        
        # Aplicar negrita a la celda "Date"
        cell = ws['B1']
        cell.font = Font(bold=True)
        
        # Aplicar negrita a toda la columna A
        for cell in ws['A']:
            cell.font = Font(bold=True)
        
        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=situation_caisse.xlsx'
        return response
    
    return render(request, 'gestion_immeuble_app/la_lista_situation_caisse.html', context)

def liste_formulaire_p_concierge(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_formulaire_p_concierge = FormulairePConcierge.objects.all()
    
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_formulaire_p_concierge, 5)
    page = request.GET.get('page')
    tous_les_formulaire_p_concierge = p.get_page(page)
    nums = "a" * tous_les_formulaire_p_concierge.paginator.num_pages
    
    print("hola : " + str(tous_les_formulaire_p_concierge.paginator.num_pages))
    
    context = {'la_lista_formulaire_p_concierge': la_lista_formulaire_p_concierge, 'tous_les_formulaire_p_concierge': tous_les_formulaire_p_concierge, 'nums': nums, 'name':name,}
    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = datetime.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulaire Concierge"
        
        # Crear una nueva hoja para la tabla
        #ws2 = wb.create_sheet(title="Formulaire Concierge")

        # Agregar los encabezados de la tabla
        headers = ['Date', 'Nom', 'Prenom', 'Montant', 'Mois', 'Année']
        #if request.user.is_superuser:
            #headers.extend(['Imprimer pdf', 'Actualiser', 'Eliminer'])
        ws.append(headers)

        # Aplicar negrita a los encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Agregar los datos de la tabla
        for item in la_lista_formulaire_p_concierge:
            row = [
                item.date.strftime('%Y-%m-%d'),
                item.nom,
                item.prenom,
                item.montant,
                item.mois,
                item.annee
            ]
            #if request.user.is_superuser:
                #row.extend(['Imprimer', 'Actualiser', 'Eliminer'])
            ws.append(row)
        
        # Crear una respuesta HTTP con el archivo Excel
        filename = f"lista_paie_concierge_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=lista_paie_concierge.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    
    return render(request, 'gestion_immeuble_app/la_lista_formulaire_p_concierge.html', context)


def enregistrer_formulaire_p_concierge_view(request):
    if request.method == 'POST':
        form_p_concierge = EnregistrerFormulairePConciergeForm(request.POST, request.FILES)
        if form_p_concierge.is_valid():
            print("Formulario P Concierge válido")
            instance_p_concierge = form_p_concierge.save(commit=False)  # Guardar el formulario sin commit para poder modificar datos antes de guardarlo completamente
            nom = instance_p_concierge.nom
            prenom = instance_p_concierge.prenom
            date = instance_p_concierge.date
            montant = instance_p_concierge.montant
            mois = instance_p_concierge.mois
            annee = instance_p_concierge.annee
            charge_id = form_p_concierge.cleaned_data.get('charge_id')  # Ejemplo de cómo obtener charge_id
            
            # Crear instancia del otro formulario y guardarla
            form_charge = EnregistrerFormulaireChargeForm({
                'charge': charge_id,  # Ejemplo de cómo podrías asignar valores al formulario de carga
                'du': date,  # Ejemplo, ajusta según tus campos
                'au': date,  # Ejemplo, ajusta según tus campos
                'date': date,  # Ejemplo, ajusta según tus campos
                'montant': montant,  # Ejemplo, ajusta según tus campos
                'image_charge': None  # Ajusta o elimina dependiendo de tus necesidades
            })
            if form_charge.is_valid():
                print("Formulario Charge válido")
                instance_charge = form_charge.save()  # Guardar el formulario de carga

                # Envío de correo electrónico
                email_message = EmailMessage(
                    subject=f'Contact Form: {nom} - {prenom} - {montant}',
                    body=f'Nom : {nom}\nPrenom: {prenom}\nDate: {date}\nMotif mois: {mois}\nMotif année: {annee}',
                    from_email=settings.EMAIL_HOST_USER,
                    to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                    reply_to=['msebti2@gmail.com']
                )
                email_message.send(fail_silently=False)

                form_p_concierge.save()  # Guardar el formulario original también si es necesario
                return redirect('liste_formulaire_p_concierge')  # Redirigir después de guardar
            else:
                print("Errores en el formulario Charge:", form_charge.errors)
                # Manejo de errores si el formulario de carga no es válido
                # Puedes retornar errores o realizar otras acciones aquí
                pass
    else:
        print("Errores en el formulario P Concierge:")
        form_p_concierge = EnregistrerFormulairePConciergeForm()

    return render(request, 'enregistrer_formulaire_p_concierge.html', {'form': form_p_concierge})


def enregistrer_formulaire_p_concierge_view2(request):
    if request.method == 'POST':
        form = EnregistrerFormulairePConciergeForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            nom = instance.nom
            prenom = instance.prenom
           
            date = instance.date
            montant=instance.montant
            
            mois = instance.mois
            annee = instance.annee
            
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {nom} - {prenom} - {montant}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Nom : {nom}\nPrenom: {prenom}\nDate: {date}\nMotif mois: {mois}\nMotif année: {annee}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            #if image_charge:
                #mime_type, _ = mimetypes.guess_type(image_charge.path)
                #email_message.attach(image_charge.name, image_charge.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)
            
            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_formulaire_p_concierge')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerFormulairePConciergeForm()
    return render(request, 'enregistrer_formulaire_p_concierge.html', {'form': form})

def actualiser_formulaire_p_concierge(request, id_formulaire_p_concierge):
    formulaire_p_concierge = FormulairePConcierge.objects.get(pk=id_formulaire_p_concierge)
    form = EnregistrerFormulairePConciergeForm(request.POST or None, request.FILES or None,  instance=formulaire_p_concierge)
    if form.is_valid():
        form.save()
        return redirect('liste_formulaire_p_concierge')
    context = {'formulaire_p_concierge': formulaire_p_concierge, 'form': form}
    return render(request, 'gestion_immeuble_app/actualizer_formulaire_p_concierge.html', context)

def eliminer_formulaire_p_concierge(request, id_formulaire_p_concierge):
    formulaire_p_concierge = get_object_or_404(FormulairePConcierge, id=id_formulaire_p_concierge)
    formulaire_p_concierge.delete()
    messages.success(request, "Le formulaire P Concierge a été eliminer correctement.")
    return redirect('liste_formulaire_p_concierge')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

def formulaire_cotization_pdf(request):
    # Crear buffer de flujo de bytes
    buf = io.BytesIO()
    # Crear canvas
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    # Crear objeto de texto
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    la_lista_formulaire_cotization = FormulaireCotization.objects.all()

    lines = []
    for formulaire_cotization in la_lista_formulaire_cotization:
        lines.append(formulaire_cotization.date.strftime('%Y-%m-%d'))  # Convertir a cadena
        lines.append(str(formulaire_cotization.liste_proprietaire.proprietaire_locataire))
        lines.append(formulaire_cotization.nom)
        lines.append(formulaire_cotization.prenom)
        lines.append(str(formulaire_cotization.aptNum))
        lines.append(str(formulaire_cotization.montant))  # Convertir a cadena
        lines.append(str(formulaire_cotization.regle))
        lines.append(str(formulaire_cotization.motif_mois))
        lines.append(str(formulaire_cotization.motif_annee))  # Convertir a cadena
        lines.append(str(formulaire_cotization.frais_sindic))  # Convertir a cadena
        lines.append(str(formulaire_cotization.frais_sindic_manual))  # Convertir a cadena
        lines.append('===============================================')

    for line in lines:
        textob.textLine(line)

    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='formulaire_cotization.pdf')

def formulaire_cotization_pdf2(request,id_formulaire_cotization):
    # Crear buffer de flujo de bytes
    buf = io.BytesIO()
    # Crear canvas
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    # Crear objeto de texto
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    la_lista_formulaire_cotization = FormulaireCotization.objects.all()
    formulaire_cotization = get_object_or_404(FormulaireCotization, id=id_formulaire_cotization)
    
    lines = []
    #for formulaire_cotization in la_lista_formulaire_cotization:
    lines.append(formulaire_cotization.date.strftime('%Y-%m-%d'))  # Convertir a cadena
    lines.append(str(formulaire_cotization.liste_proprietaire.proprietaire_locataire))
    lines.append(formulaire_cotization.nom)
    lines.append(formulaire_cotization.prenom)
    lines.append(str(formulaire_cotization.aptNum))
    lines.append(str(formulaire_cotization.montant))  # Convertir a cadena
    lines.append(str(formulaire_cotization.regle))
    lines.append(str(formulaire_cotization.motif_mois))
    lines.append(str(formulaire_cotization.motif_annee))  # Convertir a cadena
    lines.append(str(formulaire_cotization.frais_sindic))  # Convertir a cadena
    lines.append(str(formulaire_cotization.frais_sindic_manual))  # Convertir a cadena
    lines.append('===============================================')

    for line in lines:
        textob.textLine(line)

    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='formulaire_cotization_' + str(id_formulaire_cotization)  +'.pdf')


def formulaire_cotization_pdf3(request, id_formulaire_cotization):
    # Crear buffer de flujo de bytes
    buf = io.BytesIO()
    # Crear canvas
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    
    # Obtener el objeto por ID
    formulaire_cotization = get_object_or_404(FormulaireCotization, id=id_formulaire_cotization)
    
    # Configurar el texto
    c.setFont("Helvetica", 12)
    
    # Agregar el contenido basado en la imagen proporcionada
    c.drawString(1 * inch, 10 * inch, "Code")
    c.drawString(6 * inch, 10 * inch, "Bonan")
    
    c.setFont("Helvetica-Bold", 14)
    c.drawString(3 * inch, 9.5 * inch, "REÇU")
    
    c.setFont("Helvetica", 12)
    c.drawString(1 * inch, 9 * inch, f"Reçu de: {formulaire_cotization.nom} {formulaire_cotization.prenom}")
    c.drawString(1 * inch, 8.5 * inch, f"Apt N°: {formulaire_cotization.aptNum}")
    c.drawString(1 * inch, 8 * inch, f"En qualité de: {formulaire_cotization.liste_proprietaire.proprietaire_locataire}")
    c.drawString(1 * inch, 7.5 * inch, f"La somme de: {formulaire_cotization.montant} (en toutes lettres)")
    c.drawString(1 * inch, 7 * inch, f"Pour motif: Frais syndic ({formulaire_cotization.motif_mois}/{formulaire_cotization.motif_annee})")
    c.drawString(1 * inch, 6.5 * inch, f"Date: {formulaire_cotization.date.strftime('%Y-%m-%d')}")
    
    c.drawString(1 * inch, 5.5 * inch, "Signature Propriétaire/Locataire")
    c.drawString(6 * inch, 5.5 * inch, "Signature SYNDIC")
    
    c.showPage()
    c.save()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename=f'formulaire_cotization_{id_formulaire_cotization}.pdf')



def draw_content(c, formulaire_cotization, offset_y):
    # Obtener el último objeto creado
    ultimo_registro = FormulaireCotization.objects.latest('id')

# Acceder al ID del último registro
    ultimo_id = ultimo_registro.id
    c.setFont("Helvetica", 12)
    print("El formulaire_cotization.id : " + str(formulaire_cotization.id))
    # Ajustar las posiciones Y según el offset
    if formulaire_cotization.codeFormCotiz is not None:
        new_id = formulaire_cotization.codeFormCotiz
        c.drawString(1 * inch, 6.5 * inch - offset_y, f"{str(new_id)}")
    else:
        c.drawString(1 * inch, 6.5 * inch - offset_y, f"{str(ultimo_id)}")
    #c.drawString(1 * inch, 6.5 * inch - offset_y, f"{str(formulaire_cotization.id + 2205) } " )
    c.drawString(6 * inch, 6.5 * inch - offset_y, f"{formulaire_cotization.montant}")
    
    c.setFont("Helvetica-Bold", 14)
    c.drawString(3 * inch, 6.9 * inch - offset_y, "REÇU")
    
    c.setFont("Helvetica", 12)
    c.drawString(1 * inch, 7.1 * inch - offset_y, f"Reçu de: {formulaire_cotization.nom} {formulaire_cotization.prenom}")
    c.drawString(1 * inch, 7.3 * inch - offset_y, f"Apt N°: {formulaire_cotization.aptNum}")
    c.drawString(1 * inch, 7.5 * inch - offset_y, f"En qualité de: {formulaire_cotization.liste_proprietaire.proprietaire_locataire}")
    c.drawString(1 * inch, 7.7 * inch - offset_y, f"La somme de: {formulaire_cotization.montant} (en toutes lettres)")
    c.drawString(1 * inch, 7.9 * inch - offset_y, f"Pour motif: {formulaire_cotization.frais_sindic}/{formulaire_cotization.frais_sindic_manual} ({formulaire_cotization.motif_mois}/{formulaire_cotization.motif_annee})")
    c.drawString(3 * inch, 8.3 * inch - offset_y, f"Rabat le: {formulaire_cotization.date.strftime('%Y-%m-%d')}")
    
    c.drawString(1 * inch, 8.7 * inch - offset_y, f"Signature ")
    c.drawString(1 * inch, 9.1 * inch - offset_y, f"{formulaire_cotization.liste_proprietaire.proprietaire_locataire}")
    c.drawString(6 * inch, 8.7 * inch - offset_y, "Signature")
    c.drawString(6 * inch, 9.1 * inch - offset_y, "SYNDIC")

def formulaire_cotization_pdf(request, id_formulaire_cotization):
    # Crear buffer de flujo de bytes
    buf = io.BytesIO()
    # Crear canvas
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    
    # Obtener el objeto por ID
    formulaire_cotization = get_object_or_404(FormulaireCotization, id=id_formulaire_cotization)
    
    # Dibujar el contenido en la parte superior
    draw_content(c, formulaire_cotization, offset_y=1)
    
    # Dibujar línea separadora
    c.line(0.5 * inch, 5 * inch, 7.5 * inch, 5 * inch)
    
    # Dibujar el contenido en la parte inferior
    draw_content(c, formulaire_cotization, offset_y=6 * inch)
    
    c.showPage()
    c.save()
    buf.seek(0)
    
    return FileResponse(buf, as_attachment=True, filename=f'formulaire_cotization_{id_formulaire_cotization}.pdf')


def generate_temp_pdf(request):
    if request.method == 'POST':
        form = EnregistrerFormulaireCotizationForm(request.POST)
        if form.is_valid():
            # Crear un objeto temporal de FormulaireCotization
            formulaire_cotization = form.save(commit=False)

            buf = io.BytesIO()
            c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
            
            # Dibujar el contenido en la parte superior
            draw_content(c, formulaire_cotization, offset_y=0)
            
            # Dibujar línea separadora
            c.line(0.5 * inch, 5 * inch, 7.5 * inch, 5 * inch)
            
            # Dibujar el contenido en la parte inferior
            draw_content(c, formulaire_cotization, offset_y=6 * inch)
            
            c.showPage()
            c.save()
            buf.seek(0)

            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            with open(temp_file.name, 'wb') as f:
                f.write(buf.getvalue())

            response = {
                'pdf_url': request.build_absolute_uri(reverse('serve_temp_pdf', args=[os.path.basename(temp_file.name)]))
            }

            return JsonResponse(response)
    
    return JsonResponse({'error': 'Invalid data'}, status=400)

def serve_temp_pdf(request, filename):
    filepath = os.path.join(tempfile.gettempdir(), filename)
    return FileResponse(open(filepath, 'rb'), content_type='application/pdf')


def formulaire_p_concierge_pdf(request, id_formulaire_p_concierge):
    # Crear buffer de flujo de bytes
    buf = io.BytesIO()
    # Crear canvas
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    
    # Obtener el objeto por ID
    formulaire_p_concierge = get_object_or_404(FormulairePConcierge, id=id_formulaire_p_concierge)
    
    # Dibujar el contenido en la parte superior
    draw_content_p_concierge(c, formulaire_p_concierge, offset_y=1)
    
    # Dibujar línea separadora
    c.line(0.5 * inch, 5 * inch, 7.5 * inch, 5 * inch)
    
    # Dibujar el contenido en la parte inferior
    draw_content_p_concierge(c, formulaire_p_concierge, offset_y=6 * inch)
    
    c.showPage()
    c.save()
    buf.seek(0)
    
    return FileResponse(buf, as_attachment=True, filename=f'formulaire_p_concierge_{id_formulaire_p_concierge}.pdf')

def draw_content_p_concierge2(c, formulaire_p_concierge, offset_y):
    # Obtener el último objeto creado
    ultimo_registro = FormulairePConcierge.objects.latest('id')

# Acceder al ID del último registro
    ultimo_id = ultimo_registro.id
    c.setFont("Helvetica", 12)
    print("El formulaire_p_concierge.id : " + str(formulaire_p_concierge.id))
    
    c.setFont("Helvetica-Bold", 14)
    c.drawString(1 * inch, 6.5 * inch - offset_y, f"RABAT LE: {formulaire_p_concierge.date}")
    
    c.setFont("Helvetica", 12)
    c.drawString(1 * inch, 7.1 * inch - offset_y, f"Je soussigné, : {formulaire_p_concierge.nom}  {formulaire_p_concierge.prenom}, en qualité du concierge de l'immeuble 41 rue oued sebou. déclare avoir reçu la somme de {formulaire_p_concierge.montant} dh correspondante á la paie du mois de {formulaire_p_concierge.mois} {formulaire_p_concierge.annee}. ")
        
    c.drawString(6 * inch, 8.7 * inch - offset_y, "Signature")
    

def draw_content_p_concierge(c, formulaire_p_concierge, offset_y):
    # Obtener el último objeto creado
    ultimo_registro = FormulairePConcierge.objects.latest('id')
    ultimo_id = ultimo_registro.id

    c.setFont("Helvetica", 12)
    print("El formulaire_p_concierge.id : " + str(formulaire_p_concierge.id))
    
    c.setFont("Helvetica-Bold", 14)
    #c.drawString(1 * inch, 6.5 * inch - offset_y, f"RABAT LE: {formulaire_p_concierge.date}")
    date_str = formulaire_p_concierge.date.strftime("%Y-%m-%d")  # Formatear la fecha
    c.drawString(1 * inch, 6.5 * inch - offset_y, f"RABAT LE: {date_str}")
    
    c.setFont("Helvetica", 12)
    texto = (f"Je soussigné, : '{formulaire_p_concierge.nom} {formulaire_p_concierge.prenom}', en qualité "
             f"du concierge de l'immeuble 41 rue oued sebou. déclare avoir reçu la somme de "
             f"{formulaire_p_concierge.montant} dh correspondante á la paie du mois de "
             f"{formulaire_p_concierge.mois} {formulaire_p_concierge.annee}.")
    
    # Dividir el texto en líneas
    max_width = 6 * inch  # Ancho máximo del texto en el PDF
    lines = split_text_into_lines(texto, max_width, c)
    
    y = 7.1 * inch - offset_y
    for line in lines:
        c.drawString(1 * inch, y, line)
        y += 0.2 * inch  # Espacio entre líneas
    
    c.drawString(6 * inch, 8.7 * inch - offset_y, "Signature")

def split_text_into_lines(texto, max_width, canvas):
    """Divide el texto en líneas para que se ajuste dentro del ancho máximo."""
    words = texto.split()
    lines = []
    line = ""
    for word in words:
        if canvas.stringWidth(line + word + " ") < max_width:
            line += word + " "
        else:
            lines.append(line)
            line = word + " "
    lines.append(line)  # Añadir la última línea
    return lines


def generate_p_concierge_temp_pdf(request):
    if request.method == 'POST':
        form = EnregistrerFormulairePConciergeForm(request.POST)
        if form.is_valid():
            # Crear un objeto temporal de FormulaireCotization
            formulaire_p_concierge = form.save(commit=False)

            buf = io.BytesIO()
            c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
            
            # Dibujar el contenido en la parte superior
            draw_content_p_concierge(c, formulaire_p_concierge, offset_y=0)
            
            # Dibujar línea separadora
            c.line(0.5 * inch, 5 * inch, 7.5 * inch, 5 * inch)
            
            # Dibujar el contenido en la parte inferior
            draw_content_p_concierge(c, formulaire_p_concierge, offset_y=6 * inch)
            
            c.showPage()
            c.save()
            buf.seek(0)

            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            with open(temp_file.name, 'wb') as f:
                f.write(buf.getvalue())

            response = {
                'pdf_url': request.build_absolute_uri(reverse('serve_p_concierge_temp_pdf', args=[os.path.basename(temp_file.name)]))
            }

            return JsonResponse(response)
    
    return JsonResponse({'error': 'Invalid data'}, status=400)

def serve_p_concierge_temp_pdf(request, filename):
    filepath = os.path.join(tempfile.gettempdir(), filename)
    return FileResponse(open(filepath, 'rb'), content_type='application/pdf')


def backup_database(request):
    try:
         # Obtener la fecha y hora actual
        now = datetime.now()
        
        # Formatear la fecha y hora como parte del nombre del archivo
        timestamp = now.strftime('%Y-%m-%d_%H-%M-%S')
        
        # Define la ruta donde se almacenará la copia de seguridad dentro del proyecto
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # Nombre del archivo de copia de seguridad
        #backup_file_name = 'backup.json'
        backup_file_name = f'backup_{timestamp}.json'
        
        # Ruta completa del archivo de copia de seguridad
        backup_file_path = os.path.join(backup_dir, backup_file_name)
        
        # Genera la copia de seguridad
        #management.call_command('dumpdata', stdout=open(backup_file_path, 'w'))
        management.call_command('dumpdata', format='json', indent=2, output=backup_file_path)
        
        
        # Enviar el archivo por correo electrónico
        subject = 'Copia de seguridad de base de datos'
        message = 'Adjunto encontrarás la copia de seguridad de la base de datos.'
        email_message = EmailMessage(
                subject=f'Contact Form: {subject}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'{message}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            #if image_charge:
                #mime_type, _ = mimetypes.guess_type(image_charge.path)
                #email_message.attach(image_charge.name, image_charge.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)
            
            # Enviar el email
       
        
        # Adjuntar el archivo de copia de seguridad al correo
        email_message.attach_file(backup_file_path)
        
        #email_message.send(fail_silently=False)
        # Enviar el correo electrónico
        email_message.send()
        
        #return HttpResponse("Copia de seguridad enviada por correo electrónico correctamente.")
        return redirect('home')
    
    
    except Exception as e:
        return HttpResponse(f"Error al crear o enviar la copia de seguridad: {str(e)}")
